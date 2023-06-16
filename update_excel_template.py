def update_template(stock_summary, input_rawdata):
    import pandas as pd
    from openpyxl import load_workbook
    import shutil
    symbol = stock_summary.index[0]
    folder = r'C:\Users\OHAD\Google Drive\Python\StocksScraper'
    shutil.copy("ValueInvestmentPythonTemplate.xlsx", f"{symbol}"+"_Valuation.xlsx")
    file_path =  folder+'\\' + f"{symbol}"+"_Valuation.xlsx"
    wb = load_workbook(file_path)
    sheet_name = symbol + " Valuation"
    ws = wb['TempSheetName']
    ws.title = sheet_name

    # Part A - Key and general data # TODO:  summary from Yahoo.com, try to create links to macrotrends and simplywallst and finviz
    ws['A1'] = symbol # TODO:  add company name
    link_10K = "https://www.sec.gov/cgi-bin/browse-edgar?action=getcompany&CIK="+ f"{symbol}"+ "&type=10-k"
    link_10q = "https://www.sec.gov/cgi-bin/browse-edgar?action=getcompany&CIK=" + f"{symbol}" + "&type=10-q"
    link_stockrow = "https://stockrow.com/"+f"{symbol}"
    link_gurufocus = "https://www.gurufocus.com/stock/" + f"{symbol}" +"/summary"
    link_yahoo = "https://finance.yahoo.com/quote/"+ f"{symbol}"+ "?p=" + f"{symbol}"
    link_zacks = "https://www.zacks.com/stock/quote/" + f"{symbol}"

    ws['A4'].hyperlink = link_10K
    ws['A4'].value = '10-K'
    ws['A4'].style = "Hyperlink"

    ws['A5'].hyperlink = link_10q
    ws['A5'].value = '10-Q'
    ws['A5'].style = "Hyperlink"

    ws['A6'].hyperlink = link_stockrow
    ws['A6'].value = 'Stockrow'
    ws['A6'].style = "Hyperlink"

    ws['A7'].hyperlink = link_gurufocus
    ws['A7'].value = 'GuruFocus'
    ws['A7'].style = "Hyperlink"

    ws['A8'].hyperlink = link_yahoo
    ws['A8'].value = 'Yahoo'
    ws['A8'].style = "Hyperlink"

    ws['A8'].hyperlink = link_zacks
    ws['A8'].value = "Zack's"
    ws['A8'].style = "Hyperlink"

    ws['C8'] = stock_summary["Zacks Expected GR [%]"][0]
    ws['C9'] = stock_summary["Market Cap [B$]"][0]
    cash_table = input_rawdata.loc[['Cash Equiv.', 'Long Term Debt']]
    ws['B45'] = cash_table.loc["Cash Equiv."][0]
    ws['B46'] = cash_table.loc["Long Term Debt"][0]
    ws['B54'] = input_rawdata.loc["Capital Expenditures"][0]
    ws['B55'] = input_rawdata.loc["Income Tax Provisions"][0]
    wb.save(file_path)

    # Part B - Raw Data: Years range B30 to J30 (9 columns)
    raw_data_table = input_rawdata.loc[['Shareholders Equity', 'Dividends', 'Buybacks', 'Revenue', 'Operating Income',
                                        'Net Income','EPS','Free Cash Flow']]  # Select Rows by Index Label List
    raw_data_table = raw_data_table.iloc[:, 0:9]

    book = load_workbook(file_path)
    writer = pd.ExcelWriter(file_path, engine='openpyxl')
    writer.book = book
    raw_data_table.to_excel(writer, sheet_name=sheet_name, header=True,index=False, index_label=None, startrow=29, startcol=1, engine=None, merge_cells=True)

    # Part C YoY growth rates: Years range B30 to  I30 (8 columns) - this part is done by Excel formula

    # Part D other metrics (9 columns)
    metrics_table =  input_rawdata.loc[['ROIC', 'Net Profit Margin']]
    metrics_table = metrics_table.iloc[:,0:9]
    metrics_table.to_excel(writer, sheet_name=sheet_name, header=True, index=False, index_label=None, startrow=40,startcol=1, engine=None, merge_cells=True)
    # cash_table = cash_table.iloc[:,0:3]
    # cash_table.to_excel(writer, sheet_name=sheet_name, header=False, index=False, index_label=None, startrow=43,startcol=1, engine=None, merge_cells=True)

    # Part E - Summary #TODO: not working, try to get company data from some website

    # Part F - Owners Earnings and Sloan Ratio
    owners_earning_table = input_rawdata.loc[['Cash Flow from Operations', 'Cash Flow from Investments',"Total Assets"]]
    owners_earning_table = owners_earning_table.iloc[:, 0:9]
    capex = [] #TODO
    income_tax = []  # TODO

    owners_earning_table.to_excel(writer, sheet_name=sheet_name, header=True, index=False, index_label=None, startrow=49,
                        startcol=1, engine=None, merge_cells=True)

    # writer.save()
    writer.close()
    # subprocess.Popen(r'explorer C:\Users\OHAD\Google Drive\Python\StocksScraper')
    # subprocess.Popen(r'explorer '+ file_path)