import time

import numpy as np
import pandas as pd
from zacks import import_zacks_earnings
from datetime import datetime

class STOCK:
    def __init__(self):
        self.name = None
        self.ticker = None
        self.description_text = None
        self.industry = None
        self.raw_data = None
        self.zacks_expected_GR = None
        self.market_cap = None
        self.stock_price = None
        self.year_high = None
        self.year_low = None
        self.raw_data_filename  = None
        self.url = None
        self.valuation_filename = None
        self.current_discount = None

def update_template(stock):
    import pandas as pd
    from openpyxl import load_workbook
    import shutil
    # try:
    current_date = datetime.today().strftime('%d_%m_%y')
    symbol = stock.ticker.upper()
    folder = r'C:\Users\OHAD\Google Drive\QualityInvetments\Valuations'
    file_path =  folder + '\\' + f"{symbol}"+"_Valuation_" + f"{current_date}"+ ".xlsx"
    shutil.copy(folder+"\ValueInvestmentPythonTemplate.xlsx",file_path)
    wb = load_workbook(file_path)
    # sheet_name = symbol + " Valuation"
    sheet_name = "Valuation"
    ws = wb['Valuation']
    ws.title = sheet_name

    # Part A - Key and general data # TODO:  summary from Yahoo.com, try to create links to macrotrends and simplywallst and finviz
    ws['A1'] = symbol # TODO:  add company name
    ws['C1'] =  current_date
    link_10K = "https://www.sec.gov/cgi-bin/browse-edgar?action=getcompany&CIK="+ f"{symbol}"+ "&type=10-k"
    link_10q = "https://www.sec.gov/cgi-bin/browse-edgar?action=getcompany&CIK=" + f"{symbol}" + "&type=10-q"
    link_stockrow = "https://stockrow.com/"+f"{symbol}"
    link_gurufocus = "https://www.gurufocus.com/stock/" + f"{symbol}" +"/summary"
    link_yahoo = "https://finance.yahoo.com/quote/"+ f"{symbol}"+ "?p=" + f"{symbol}"
    link_zacks = "https://www.zacks.com/stock/quote/" + f"{symbol}"
    link_seekingalpha = "https://seekingalpha.com/symbol/" + f"{symbol}"
    link_ycharts = "https://ycharts.com/companies/"+f"{symbol}"+"/pe_ratio"

    ws['A4'].hyperlink = link_10K
    ws['A4'].value = '10-K'
    ws['A4'].style = "Hyperlink"

    ws['A5'].hyperlink = link_10q
    ws['A5'].value = '10-Q'
    ws['A5'].style = "Hyperlink"

    ws['A6'].hyperlink = link_stockrow
    ws['A6'].value = 'Stockrow'
    ws['A6'].style = "Hyperlink"

    ws['A7'].hyperlink = link_gurufocus
    ws['A7'].value = 'GuruFocus'
    ws['A7'].style = "Hyperlink"

    ws['A8'].hyperlink = link_yahoo
    ws['A8'].value = 'Yahoo'
    ws['A8'].style = "Hyperlink"

    ws['A9'].hyperlink = link_zacks
    ws['A9'].value = "Zack's"
    ws['A9'].style = "Hyperlink"

    ws['A10'].hyperlink = link_seekingalpha
    ws['A10'].value = "Seeking Alpha"
    ws['A10'].style = "Hyperlink"

    ws['A14'].hyperlink = link_ycharts
    ws['A14'].value = "Y charts"
    ws['A14'].style = "Hyperlink"

    ws['C8'] = stock.zacks_expected_GR
    ws['C9'] = stock.market_cap
    ws['I12'] = stock.stock_price
    ws['C10'] = (stock.stock_price- stock.year_high) / stock.year_high
    ws['C11'] = (stock.stock_price - stock.year_low) / stock.year_low
    # cash_table = input_rawdata.loc[['Cash Equiv.', 'Long Term Debt']]
    # ws['B54'] = stock.raw_data.loc["Cash and Short Term Investments"][0]
    # ws['B55'] = stock.raw_data.loc["Long Term Debt (Total)"][0]
    ws['B65'] = stock.raw_data.loc["Shares (Common)"][0]
    ws['B63'] =  stock.raw_data.loc["Capital Expenditures"][0]
    try:
        ws['B64'] = stock.raw_data.loc["Income Tax Provision"][0]
    except:
        print('No tax provisions')
        ws['B64'] = 0
    wb.save(file_path)

    # Part B - Raw Data: Years range B30 to J30 (9 columns)
    # TODO: check if "Equity Repurchase (Common, Net)" == buybacks?
    raw_data_table = stock.raw_data.loc[['Shareholders Equity (Total)', 'Dividends Paid (Common)', 'Equity Repurchase (Common, Net)', 'Revenue', 'Operating Income',
                                        'Net Income','EPS (Diluted)','Free Cash Flow']]  # Select Rows by Index Label List
    raw_data_table = raw_data_table.iloc[:, 0:9]
    raw_data_table.loc['Equity Repurchase (Common, Net)'] = -1 * raw_data_table.loc['Equity Repurchase (Common, Net)']

    book = load_workbook(file_path)
    writer = pd.ExcelWriter(file_path, engine='openpyxl')
    writer.book = book
    raw_data_table.to_excel(writer, sheet_name=sheet_name, header=True,index=False, index_label=None, startrow=37, startcol=1, engine=None, merge_cells=True)
    PE_tbl= stock.Qratios.loc["P/E ratio"].T
    PE_tbl.to_excel(writer, sheet_name=sheet_name, header=True, index=False, index_label=None, startrow=29, startcol=11, engine=None, merge_cells=True)
    # Part C YoY growth rates: Years range B30 to  I30 (8 columns) - this part is done by Excel formula

    # Part D other metrics (9 columns)
    metrics_table =  stock.raw_data.loc[['ROIC', 'Net Profit Margin', "Cash and Short Term Investments", "Long Term Debt (Total)"]]

    try:
        inventory_row =  stock.raw_data.loc[['Inventory']]
    except Exception as e:
        inventory_row = np.empty((1,metrics_table.shape[1]))
        inventory_row = pd.DataFrame([])
        print(e)
    # metrics_table = metrics_table.iloc[:,0:9]
    metrics_table = pd.concat([metrics_table, inventory_row], axis = 0)
    metrics_table = metrics_table.drop_duplicates()
    # metrics_table = metrics_table[~metrics_table.index.duplicated(keep='first')]  # Remove duplicate rows
    metrics_table.to_excel(writer, sheet_name=sheet_name, header=True, index=False, index_label=None, startrow=48,startcol=1, engine=None, merge_cells=True)

    # cash_table = cash_table.iloc[:,0:3]
    # cash_table.to_excel(writer, sheet_name=sheet_name, header=False, index=False, index_label=None, startrow=43,startcol=1, engine=None, merge_cells=True)

    # Part E - Summary #TODO: not working, try to get company data from some website

    # Part F - Owners Earnings and Sloan Ratio
    # TODO: check if 'Cash Flow from Operations' == 'Income from Continuous Operations'
    owners_earning_table = stock.raw_data.loc[['Income from Continuous Operations', 'Investing cash flow',"Total Assets"]]
    owners_earning_table = owners_earning_table.iloc[:, 0:9]
    owners_earning_table.to_excel(writer, sheet_name=sheet_name, header=True, index=False, index_label=None, startrow=58, startcol=1, engine=None, merge_cells=True)
    writer.close()
    # subprocess.Popen(r'explorer C:\Users\OHAD\Google Drive\Python\StocksScraper')
    # subprocess.Popen(r'explorer '+ file_path)
    return file_path
    # except Exception as e:
    #     print(f'####### Template updating failed: {e} ####### ')


def get_data_from_stock_excel(filename):
    fail_counter = 0
    MAX_FAILS = 5
    while fail_counter<MAX_FAILS:
        try:
            stock_income_statement_df = pd.read_excel(filename, sheet_name="Income Statement, A")
            break
        except Exception as e:
            fail_counter +=1
            print(f"Failed on try {fail_counter}: {e}")
            time.sleep(2)
    if fail_counter == MAX_FAILS:
        print("****** Failed to perform valuation! ***** ")

    stock_balance_sheet_df = pd.read_excel(filename, sheet_name="Balance Sheet, A")
    stock_metrics_df = pd.read_excel(filename, sheet_name="Metrics Ratios, A")
    stock_cashflow_df = pd.read_excel(filename, sheet_name="Cash Flow, A")
    stock_Qratios_df = pd.read_excel(filename, sheet_name="Metrics Ratios, Q", index_col=0)

    stock_data_df = pd.concat([stock_income_statement_df, stock_balance_sheet_df, stock_cashflow_df, stock_metrics_df])
    stock_data_df = stock_data_df.set_index('Unnamed: 0')
    if stock_data_df.shape[0] != stock_income_statement_df.shape[0] + stock_balance_sheet_df.shape[0] + stock_metrics_df.shape[0] + stock_cashflow_df.shape[0]:
        print("Something wrong with shape of stock_data_df")

    return stock_data_df, stock_Qratios_df

def get_stock_data(ticker, filename):
    stock = STOCK()
    stock.ticker = ticker.upper()
    zacks_result = import_zacks_earnings(stock.ticker)
    stock.zacks_expected_GR = zacks_result.Exp_EPS_GR
    stock.stock_price = zacks_result.stock_price
    stock.year_low = zacks_result.TTM_low
    stock.year_high = zacks_result.TTM_high
    stock.market_cap = zacks_result.market_cap
    stock.raw_data, stock.Qratios = get_data_from_stock_excel(filename)
    return stock

if __name__ == '__main__':
    filename = r"C:\Users\OHAD\Google Drive\QualityInvetments\stockrow_excels\financials_export_amat_2024_01_28_07_25_15.xlsx"
    ticker = filename.split("financials_export_")[1]
    stock = get_stock_data(ticker, filename)
    stock.ticker = filename.split("financials_export_")[1]
    result = update_template(stock)
    print(f"Done making valuation for {stock.ticker}")